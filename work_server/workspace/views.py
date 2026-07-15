# ---------- БЛОК ИМПОРТОВ ----------

# Импорты из Django
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch, OuterRef, Exists, Q
from django.core.exceptions import ValidationError
from django.core.cache import cache
from django.http import HttpResponseRedirect, HttpResponseForbidden, JsonResponse, HttpRequest, HttpResponse
from django.contrib.auth.models import Group, User
from django.utils import timezone
# Импорты из текущего приложения
from .models import *
from .forms import *
# Импорты из сторонних библиотек
from dateutil.relativedelta import relativedelta
import hashlib
import json
import pandas as pd
import openpyxl as xl
from fnmatch import fnmatch
import re


# ---------- БЛОК ВСПОМОГАТЕЛЬНЫХ ФУНКЦИЙ ----------


# Функции для получения стандартных состояний объектов
# Нужны для работы сервера в async режиме (при статичном получении выдаст ошибку)
def get_default_object_state():
    return ObjectState.objects.filter(name="Приостановлен").first()


def get_ready_object_state():
    return ObjectState.objects.filter(name="В сборке").first()


# Символы, используемые в именах объектов и изделий
ALPHABET = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюяabcdefghijklmnopqrstuvwxyz!@#$%^&*()-=_+"№;:?'


def check_worker_data(request: HttpRequest = None, user: User = None) -> WorkerData:
    """
    ### Описание
    Проверяет существование данных о работнике, возвращает созданную модель, если данных нет
    ### Параметры
    *требуется хотя бы один из параметров*
    - request — HTTP-запрос
    - user — пользователь
    ### Возвращаемое значение
    - Модель WorkerData, соответствующая работнику
    """
    # Проверяет, что пользователь передан (эквивалентно user != None)
    if user:
        # Проверяем наличие записи в БД
        if WorkerData.objects.filter(worker=user).exists():
            return WorkerData.objects.filter(worker=user).first()
        else:
            return WorkerData.objects.create(worker=user)
    # Если передан HTTP-запрос (request != None)
    elif request:
        # Проверяем наличие записи в БД
        if WorkerData.objects.filter(worker=request.user).exists():
            return WorkerData.objects.filter(worker=request.user).first()
        else:
            return WorkerData.objects.create(worker=request.user)
    else:
        raise KeyError("One argument required: request OR user")


def check_user_group(request: HttpRequest, group_name: str) -> bool:
    """
    ### Описание
    Проверяет принадлежность пользователя указанной группе
    ### Параметры
    - request — HTTP-запрос
    - group_name — имя группы, которой должен принадлежать пользователь
    ### Возвращаемое значение
    - True, если пользователь принадлежит группе
    - False, если пользователь не принадлежит группе
    """
    # Получаем группу по имени
    target_group = Group.objects.get(name=group_name)
    # Если группа не найдена, выдаём ошибку
    if not target_group:
        raise ValidationError(f'Группа с именем {group_name} не найдена')
    # Проверяем принадлежность пользователя группе
    if target_group in request.user.groups.all():
        return True
    else:
        return False


def update_notification(request: HttpRequest = None) -> JsonResponse | None:
    """
    ### Описание
    Проверяет наличие новых уведомлений для пользователя и возвращает их в формате JSON
    ### Параметры
    - request — HTTP-запрос
    ### Возвращаемое значение
    - JsonResponse с уведомлением, если оно есть
    - Пустой JsonResponse, если уведомлений нет
    - None, если уведомлений нет или запрос некорректен
    """
    # Проверяем, что пришёл AJAX-запрос на обновление уведомлений и пользователь авторизован
    if request.headers.get('X-Requested-With') and 'XMLNotificationUpdate' in request.headers.get('X-Requested-With') and request.user.is_authenticated:
        notification = None
        # Получаем уведомления, созданные за последние 2 минуты для группы пользователя
        notifications = Notification.objects.filter(
            recipient_group=request.user.groups.first(), created_at__gte=timezone.now()-relativedelta(minutes=2)).order_by('-created_at')
        # Если нашли уведомления, выбираем первое непрочитанное
        if notifications:
            for notify in notifications:
                if request.user not in notify.read_by.all():
                    notification = notify
                    break
        # Если есть непрочитанное уведомление, готовим его к отправке
        if notification:
            # Кешируем уведомление, чтобы не отправлять его повторно
            cache_key = f'notification_{request.user}'
            cur_hash = hashlib.md5(json.dumps(
                [notification.id, notification.title, notification.message], sort_keys=True).encode()).hexdigest()
            prev_hash = cache.get(cache_key)
            if prev_hash and prev_hash == cur_hash:
                return JsonResponse({'html': ""})
            cache.set(cache_key, cur_hash, timeout=300)
            # Формируем JSON-ответ с уведомлением (в нём заполняем шаблон и добавляем данные об уведомлении)
            data = {'html': render_to_string(
                "partials/notification.html", {'notification': notification}, request), 'message': notification.message, 'time': notification.created_at}
            # Отмечаем уведомление как прочитанное для данного пользователя
            notification.read_by.add(request.user)
            notification.save()
            return JsonResponse(data)
        else:
            return JsonResponse({'html': ""})
    return None

# --- LEGACY для работы со сводной ---

# def check_summary(data: pd.DataFrame):
#     """
#     Проверяет формат Сводной

#     data — данные для парсинга из Сводной в формате pandas.DataFrame
#     """
#     # Проверка заголовка сводной
#     if data.iloc[0, 2] != 'Сводная таблица закупаемого оборудования №':
#         raise ValidationError(
#             "Не удалось определить, является ли файл сводной")
#     obj_number = str(data.iloc[0, 3])
#     # Проверка заголовков для изделий
#     if data.iloc[2, 1] != 'Перечень изготавливаемых изделий:':
#         raise ValidationError("Не найден перечень изготавливаемых изделий")
#     if data.iloc[3, 1] != 'Зав. номер':
#         raise ValidationError("Не найдены заводские номера изделий")
#     if data.iloc[3, 2] != 'Наименование':
#         raise ValidationError("Не найдены наименования изделий")
#     if data.iloc[3, 3] != 'Кол-во':
#         raise ValidationError("Не найдены количества изготавливаемых изделий")
#     # Проверка формата и расположения изделий
#     nan_idx = 4
#     for value in data.iloc[4:, 1]:
#         if pd.notna(value):
#             nan_idx += 1
#             if obj_number not in value:
#                 return False
#         else:
#             break
#     # Проверка заголовков компонентов
#     if data.iloc[nan_idx+1, 1] != 'Перечень оборудования, закупаемого производством:':
#         raise ValidationError("Не найден перечень закупаемого оборудования")
#     if data.iloc[nan_idx+2, 2] != 'Наименование':
#         raise ValidationError("Не найдены наименования оборудования")
#     if data.iloc[nan_idx+2, 14] != 'Дефицит':
#         raise ValidationError(
#             "Не найдены данные о дефиците оборудования")
#     # Проверки закончены
#     return True


# ---------- БЛОК VIEW-ФУНКЦИЙ ----------
@login_required
def index(request: HttpRequest) -> JsonResponse | HttpResponse:
    """
    ### Описание
    Главная страница - список объектов или список доступных изделий в зависимости от группы пользователя
    ### Параметры
    - request — HTTP-запрос
    ### Возвращаемое значение
    - Заполненный шаблон главной страницы в зависимости от группы пользователя
    """
    # Если пользователь принадлежит группе Работник
    # Загружаем шаблон для работника
    if check_user_group(request, "worker"):
        # Проверка данных после авторизации.
        # В таком случае данные будут созданы в случае
        # Если они не существуют
        worker = check_worker_data(request)
        context = dict()
        # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
        notify = update_notification(request)
        if notify:
            return notify
        # Получаем список изделий, добавленных в очередь работника
        queued = CreationInstance.objects.filter(
            worker=worker, status='QUEUED').prefetch_related('product', 'part')
        # Если такие есть, получаем первое из них и весь список
        # Для взятия в работу будет доступно только первое
        if queued:
            context['queued_first'] = queued.first()
            context['queued'] = queued
        # Иначе получаем список всех доступных изделий
        else:
            ready_state_subquery = ObjectStateInstance.objects.filter(
                object=OuterRef('pk'),
                state=get_ready_object_state()
            )

            objects = Object.objects.filter(
                hidden=False,
                ready_percentage__lt=100
            ).annotate(
                is_ready=Exists(ready_state_subquery)
            ).filter(
                is_ready=True
            ).prefetch_related(
                Prefetch('product_set',
                         queryset=Product.objects.prefetch_related('part_set'))
            )

            product_ids = []
            for obj in objects:
                for product in obj.product_set.all():
                    product_ids.append(product.id)

            preloaded_instances = list(CreationInstance.objects.filter(
                Q(product_id__in=product_ids) | Q(
                    part__product_id__in=product_ids)
            ))

            preloaded_parts = list(
                Part.objects.filter(product_id__in=product_ids))

            products = []
            search_query = request.GET.get('search', '')

            for obj in objects:
                for product in obj.product_set.all():
                    ava_amount = product.get_ava_amount(
                        preloaded_instances=preloaded_instances,
                        preloaded_parts=preloaded_parts
                    )

                    has_available_parts = False
                    for part in product.part_set.all():
                        if part.get_ava_amount(preloaded_instances=preloaded_instances) > 0:
                            has_available_parts = True
                            break

                    if (ava_amount > 0 or has_available_parts) and search_query in product.get_id():
                        products.append(product)

            context['products'] = products

            if request.headers.get('X-Requested-With') == 'XMLHttpSearchRequest':
                data = {'html': render_to_string(
                    "partials/worker_products.html", context, request)}
                return JsonResponse(data)
        # Если пришёл AJAX-запрос на обновление списка изделий
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Проверяем кеш, чтобы не отправлять повторно одинаковые данные
            cache_key = f'worker_products_list_{request.user}'
            if queued:
                cur_hash = hashlib.md5(json.dumps(
                    list(queued.values('id')), sort_keys=True).encode()).hexdigest()
            else:
                prod_ids = []
                for product in products:
                    prod_ids.append(product.id)
                cur_hash = hashlib.md5(json.dumps(
                    prod_ids, sort_keys=True).encode()).hexdigest()
            prev_hash = cache.get(cache_key)
            if prev_hash and prev_hash == cur_hash:
                return JsonResponse({'html': ""})
            cache.set(cache_key, cur_hash, timeout=300)
            # Отправляем заполненный шаблон с изделиями
            data = {'html': render_to_string(
                "partials/worker_products.html", context, request)}
            return JsonResponse(data)

        # Выполняется при первичной загрузке страницы
        # Кешируем список изделий, чтобы не отправлять повторно одинаковые данные
        cache_key = f'worker_products_list_{request.user}'
        if queued:
            cur_hash = hashlib.md5(json.dumps(
                list(queued.values('id')), sort_keys=True).encode()).hexdigest()
        else:
            prod_ids = []
            for product in products:
                prod_ids.append(product.id)
            cur_hash = hashlib.md5(json.dumps(
                prod_ids, sort_keys=True).encode()).hexdigest()
        cache.set(cache_key, cur_hash, timeout=300)
        # Отправляем пользователю шаблон, заполняя его данными
        return render(request, "worker.html", context)
    # Если пользователь принадлежит группе Мастер
    # Загружаем шаблон для мастера
    elif check_user_group(request, "master"):
        # Получаем объекты, которые не скрыты
        objects = Object.objects.filter(hidden=False)
        # Получаем вопросы без ответов
        questions = Question.objects.filter(answer='')
        # Сохраняем данные в контекст для заполнения шаблона
        context = {'objects': objects, 'questions': len(questions)}
        # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
        notify = update_notification(request)
        if notify:
            return notify
        # Если пришёл AJAX-запрос на обновление списка объектов
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            data = []
            # Сохраняем информацию об объектах
            for object in objects:
                data.append(
                    f'obj: {object.id} ready: {object.get_ready_percentage()}')
            # Сохраняем информацию о вопросах
            for question in questions:
                data.append(f'question: {question.id}')
            # Проверяем кеш, чтобы не отправлять повторно одинаковые данные
            cur_hash = hashlib.md5(json.dumps(
                data, sort_keys=True).encode()).hexdigest()
            cache_key = f'master_object_list_{request.user}'
            prev_hash = cache.get(cache_key)
            if prev_hash and prev_hash == cur_hash:
                return JsonResponse({'html': ""})
            cache.set(cache_key, cur_hash, timeout=300)
            # Отправляем заполненный шаблон с объектами
            question_len = 0
            if questions:
                question_len = len(questions)
            data = {'html': render_to_string(
                'partials/objects_table.html', context, request), 'questions': question_len}
            return JsonResponse(data)
        # Выполняется при первичной загрузке страницы
        # Отправляем заполненный шаблон с объектами
        return render(request, 'master.html', context)

    # Иначе отправляем шаблон с текстом об ошибке (т.к. эта страница загрузится только для пользователя без группы)
    else:
        return render(request, 'index.html')


@login_required
def product_detail_view(request: HttpRequest, pk: int) -> JsonResponse | HttpResponse | HttpResponseRedirect:
    """
    ### Описание
    Детальный обзор изделия для взятия в работу
    ### Параметры
    - request — HTTP-запрос
    - pk — первичный ключ изделия (его id)
    ### Возвращаемое значение
    - Заполненный шаблон с формой для взятия изделия в работу
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "worker") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем информацию о выбранном изделии
    product = get_object_or_404(Product, pk=pk)
    raw_parts = Part.objects.filter(product=product)
    parts = None
    for part in raw_parts:
        if part.get_ava_amount() > 0:
            if parts is None:
                parts = [part]
            else:
                parts.append(part)
    # Заполняем варианты выбора для формы взятия изделия в работу
    # Благодаря ним работник сможет изготавливать целое изделие или его часть
    # Не выбирая количество вручную (т.к. есть значение по умолчанию)
    tmpl_choices = dict()
    choices = None
    # Значение по умолчанию, будет подобрано к первому доступному изделию/части
    def_amount = 1
    # Выбор по умолчанию - всё изделие или его первая доступная часть
    def_choice = '1'
    if product.get_ava_amount() > 0:
        def_amount = min(def_amount, product.ava_float())
        choices = [('1', 'Всё изделие')]
        tmpl_choices['1'] = str(def_amount)
    idx = 2
    if parts:
        for part in parts:
            # Если ещё не выбран вариант по умолчанию (всё изделие недоступно)
            # Выбираем первую доступную часть изделия
            if choices is None:
                def_amount = min(def_amount, part.get_ava_amount())
                def_choice = str(idx)
                choices = [(str(idx), part.name)]
                tmpl_choices[str(idx)] = str(min(1, part.get_ava_amount()))
                idx += 1
            else:
                choices.append((str(idx), part.name))
                tmpl_choices[str(idx)] = str(min(1, part.get_ava_amount()))
                idx += 1
    # Если была нажата кнопка Взять в работу (пришёл POST запрос)
    # Обрабатываем данные
    if request.method == "POST":
        # Получаем данные от формы
        form = TakeProductToWorkForm(
            request.POST, choices=choices, initial={'amount': def_amount, 'creation': def_choice})
        # Если не возникло ошибок
        if form.is_valid():
            # Получаем выбранное количество
            amount = form.cleaned_data['amount']
            choice = int(form.cleaned_data['creation'])
            # Если выбрано всё изделие
            if choice == 1:
                # Если количество превышает доступное
                # Выводим сообщение об ошибке (добавляя ошибку в форму, дальнейшая обработка произойдёт в шаблоне)
                if amount > product.get_ava_amount():
                    context = {
                        'form': form,
                        'product': product,
                        'parts': parts,
                        'choices': tmpl_choices,
                    }
                    form.add_error(
                        'amount', "Указанное количество изделий превышает допустимое")
                    return render(request, 'product_detail.html', context)
                # Иначе создаём запись о новом изделии в работе
                else:
                    # Получаем данные о работнике
                    worker_data = check_worker_data(request)
                    # Сбрасываем доступное количество изделия для последующего обновления
                    product.ava_amount = None
                    product.save()
                    # Сбрасываем доступное количество всех частей изделия для последующего обновления
                    parts = Part.objects.filter(product=product)
                    for part in parts:
                        part.ava_amount = None
                        part.save()
                    # Если запись уже есть, обновляем её (увеличилось кол-во изделий в работе)
                    wip_product = CreationInstance.objects.filter(
                        worker=worker_data, product=product, status="IN_WORK").first()
                    if wip_product:
                        wip_product.amount += amount
                        wip_product.save()
                    # Иначе создаём новую запись
                    else:
                        CreationInstance.objects.create(
                            product=product, worker=worker_data, amount=amount, status='IN_WORK', started=timezone.now())
                    # Обновляем доступное количество всех частей изделия
                    for part in parts:
                        part.get_ava_amount()
                    # Обновляем доступное количество изделия
                    product.get_ava_amount()
                    # Возвращаем пользователя на главную страницу
                    return HttpResponseRedirect('/workspace')
            # Если выбрана часть изделия
            # Обрабатываем аналогично
            else:
                selected_part = None
                idx = 2
                for part in parts:
                    if idx == choice:
                        selected_part = part
                        break
                    idx += 1
                if amount > selected_part.get_ava_amount():
                    context = {
                        'form': form,
                        'product': product,
                        'parts': parts,
                        'choices': tmpl_choices,
                    }
                    form.add_error(
                        'amount', "Указанное количество изделий превышает допустимое")
                    return render(request, 'product_detail.html', context)
                else:
                    # Получаем данные о работнике
                    worker_data = check_worker_data(request)
                    # Сбрасываем доступное количество для последующего обновления
                    selected_part.product.ava_amount = None
                    selected_part.product.save()
                    # Сбрасываем доступное количество всех частей изделия для последующего обновления
                    parts = Part.objects.filter(product=selected_part.product)
                    for part in parts:
                        part.ava_amount = None
                        part.save()
                    # Если запись уже есть, обновляем её (увеличилось кол-во частей в работе)
                    wip_part = CreationInstance.objects.filter(
                        worker=worker_data, part=selected_part).first()
                    if wip_part:
                        wip_part.amount += amount
                        wip_part.save()
                    # Иначе создаём новую запись
                    else:
                        CreationInstance.objects.create(
                            part=selected_part, worker=worker_data, amount=amount, status='IN_WORK', started=timezone.now())
                    # Обновляем доступное количество частей и изделия
                    for part in parts:
                        part.get_ava_amount()
                    selected_part.product.get_ava_amount()
                    # Возвращаем пользователя на главную страницу
                    return HttpResponseRedirect('/workspace')

    # Если пришёл другой запрос (GET), возвращаем шаблон с формой для взятия изделия в работу
    else:
        # Если пришёл AJAX-запрос на обновление данных изделия
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Проверяем кеш, чтобы не отправлять повторно одинаковые данные
            cache_key = f'product_detail_{request.user}'
            cache_data = [f'prod: {product.ava_float()}',
                          f'descr: {product.description}']
            if parts:
                for part in parts:
                    cache_data.append(f'{part.id}: {part.get_ava_amount}')
            cur_hash = hashlib.md5(json.dumps(
                cache_data, sort_keys=True).encode()).hexdigest()
            prev_hash = cache.get(cache_key)
            if prev_hash and prev_hash == cur_hash:
                return JsonResponse({'html': ""})
            cache.set(cache_key, cur_hash, timeout=300)
            # Если больше нечего изготавливать (изделие или все его части уже взяты в работу)
            # Возвращаем пользователя на главную
            if parts == None:
                if product.get_ava_amount() == 0:
                    data = {'return': True}
                    return JsonResponse(data)
            else:
                if product.get_ava_amount() == 0 and all(part.get_ava_amount() == 0 for part in parts):
                    data = {'return': True}
                    return JsonResponse(data)
            # Создаём форму с вариантами выбора
            form = TakeProductToWorkForm(choices=choices, initial={
                'amount': def_amount, 'creation': def_choice})
            # Отправляем заполненный шаблон с формой взятия в работу
            context = {
                'form': form,
                'product': product,
                'parts': parts,
                'choices': tmpl_choices,
            }
            data = {'html': render_to_string(
                "partials/product_details.html", context, request)}
            return JsonResponse(data)
        # При стандартной прогрузке страницы создаём форму с вариантами выбора
        else:
            form = TakeProductToWorkForm(choices=choices, initial={
                'amount': def_amount, 'creation': def_choice})
    # Отправляем заполненный шаблон с формой взятия в работу
    context = {
        'form': form,
        'product': product,
        'parts': parts,
        'choices': tmpl_choices,
    }
    return render(request, 'product_detail.html', context)


# Список изделий, изготавливаемых работником
@login_required
def my_products_view(request: HttpRequest) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Страница со списком изделий, которые в данный момент изготавливает работник
    ### Параметры
    - request — HTTP-запрос
    ### Возвращаемое значение
    - Заполненный шаблон со списком изделий в работе
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "worker") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем данные о работнике
    worker_data = check_worker_data(request)
    # Получаем запись о всех изделиях, выполняемых данным работником
    instances = CreationInstance.objects.filter(
        worker=worker_data, status='IN_WORK')
    # Отправляем заполненный шаблон
    context = {'instances': instances}
    return render(request, 'my_products.html', context)


@login_required
def my_product_view(request: HttpRequest, pk: int) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Детальный обзор изделия, которое в данный момент изготавливает работник
    ### Параметры
    - request — HTTP-запрос
    - pk — первичный ключ записи о изделии в работе (его id)
    ### Возвращаемое значение
    - Заполненный шаблон с деталями изделия в работе
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "worker") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем данные о работнике
    worker_data = check_worker_data(request)
    # Получаем запись о выбранном изделии
    instance = get_object_or_404(CreationInstance, pk=pk)
    # Если данные запросил пользователь, который не ведёт работу над изделием
    # Вовзращаем его на страницу с его изделиями
    if instance.worker != worker_data:
        return HttpResponseRedirect("/workspace/my_products")
    # Получаем данные о всех вопросах на странице данного изделия
    all_questions = Question.objects.filter(instance=instance)
    # Если получен POST запрос (нажата кнопка Отправить вопрос или Завершить изделие или Отменить изделие)
    if request.method == "POST":
        if 'send_question' in request.POST:
            # Получаем данные из формы
            form = EnterQuestionForm(request.POST, initial={'question': ' '})
            # Создаём новый вопрос, если не возникло ошибок
            if form.is_valid():
                question = form.cleaned_data['question']
                Question.objects.create(
                    instance=instance, quest=question)
                all_questions = Question.objects.filter(instance=instance)
        elif 'finish_product' in request.POST:
            # Получаем ссылку на объект изделия
            object: Object = instance.product.object if instance.product else instance.part.product.object
            # Сбрасываем готовность объекта для последующего обновления
            object.ready_percentage = None
            object.save()
            # Сбрасываем доступное количество изделия/частей и завершённого количества изделия
            parts = Part.objects.filter(product=instance.product)
            for part in parts:
                part.ava_amount = None
                part.save()
            product = instance.product if instance.product else instance.part.product
            product.ava_amount = None
            product.completed_amount = None
            product.save()
            # Если запись о завершённом изделии/части уже есть, обновляем её (увеличилось кол-во завершённых изделий/частей)
            completed = CreationInstance.objects.filter(
                worker=worker_data, product=instance.product, part=instance.part, status="COMPLETED").first()
            if completed:
                completed.amount += instance.amount
                completed.completed = timezone.now()
                completed.save()
                instance.delete()
            # Иначе помечаем текущее изделие/часть как завершённое
            else:
                instance.status = 'COMPLETED'
                instance.completed = timezone.now()
                instance.save()
            # Обновляем доступное кол-во частей
            for part in parts:
                part.get_ava_amount()
            # Обновляем доступное и завершённое количество изделия и готовность объекта
            product.get_ava_amount()
            product.get_completed_amount()
            object.get_ready_percentage()
            # Создаём уведомление для мастера о завершении изделия
            Notification.objects.create(recipient_group=Group.objects.get(
                name='master'), title='Завершено изделие', message=f'{worker_data.display_name} завершил работу над {instance}')
            # Возвращаем пользователя на страницу со списком его изделий
            return HttpResponseRedirect('/workspace/my_products')
        elif 'cancel_product' in request.POST:
            # Удаляем все вопросы, связанные с изделием/частью
            while all_questions:
                all_questions.first().delete()
            # Сбрасываем доступное количество изделия для последующего обновления
            product = instance.product if instance.product else instance.part.product
            product.ava_amount = None
            # Сбрасываем доступное кол-во частей для обновления
            parts = Part.objects.filter(product=instance.product)
            for part in parts:
                part.ava_amount = None
                part.save()
            # Удаляем запись об изготовлении изделия
            instance.delete()
            # Обновляем доступное кол-во частей
            for part in parts:
                part.get_ava_amount()
            # Обновляем доступное количество изделия
            product.get_ava_amount()
            # Возвращаем пользователя на страницу со списком его изделий
            return HttpResponseRedirect('/workspace/my_products')
    # Если получен другой запрос (GET)
    else:
        # Если пришёл AJAX-запрос на обновление списка вопросов
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Проверяем кеш, чтобы не отправлять повторно одинаковые данные
            cache_key = f'worker_product_{request.user}'
            cur_hash = hashlib.md5(json.dumps(list(all_questions.values(
                "id", "quest", "answer")), sort_keys=True).encode()).hexdigest()
            prev_hash = cache.get(cache_key)
            if prev_hash and prev_hash == cur_hash:
                return JsonResponse({'html': ""})
            cache.set(cache_key, cur_hash, timeout=300)
            # Отправляем заполненный шаблон с вопросами
            context = {
                'instance': instance,
                'questions': all_questions
            }
            data = {'html': render_to_string(
                "partials/questions_list.html", context, request)}
            return JsonResponse(data)
        # При стандартной прогрузке страницы создаём пустую форму для ввода вопроса
        else:
            form = EnterQuestionForm()
    # Отправляем заполненный шаблон
    context = {
        'form': form,
        'instance': instance,
        'questions': all_questions
    }
    return render(request, 'my_product.html', context)


@login_required
def object_detail_view(request: HttpRequest, pk: int) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Детальный обзор объекта для мастера
    ### Параметры
    - request — HTTP-запрос
    - pk — первичный ключ объекта (его id)
    ### Возвращаемое значение
    - Заполненный шаблон с деталями объекта
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем информацию о выбранном объекте
    object = get_object_or_404(Object, pk=pk)
    states = ObjectStateInstance.objects.filter(object=object)

    # Получаем данные обо всех изделиях данного объекта
    products = Product.objects.filter(object=object).prefetch_related(
        'object',
        'part_set',
        'creationinstance_set__worker',
        'part_set__creationinstance_set__worker'
    )
    # Создаём флаг возможности удаления объекта
    can_be_deleted = True
    for product in products:
        if product.get_ava_amount() != product.amount:
            can_be_deleted = False
            break
    # Создаём флаг готовности объекта
    ready = False
    for state in states:
        if state.state == get_ready_object_state():
            ready = True
            break

    for product in products:
        product.active_product_instances = [
            inst for inst in product.creationinstance_set.all()
            if inst.status == 'IN_WORK'
        ]

        active_part_instances = []
        for part in product.part_set.all():
            for inst in part.creationinstance_set.all():
                if inst.status == 'IN_WORK':
                    inst.part_name = part.name
                    active_part_instances.append(inst)
        product.active_part_instances = active_part_instances

        completed_instances = []
        for inst in product.creationinstance_set.all():
            if inst.status == 'COMPLETED':
                inst.item_name = "Целое изделие"
                completed_instances.append(inst)
        for part in product.part_set.all():
            for inst in part.creationinstance_set.all():
                if inst.status == 'COMPLETED':
                    inst.item_name = f"Деталь: {part.name}"
                    completed_instances.append(inst)
        product.completed_instances = completed_instances

    context = {
        'object': object,
        'states': states,
        'products': products,
        'delete': can_be_deleted,
        'ready': ready
    }
    if request.method == "POST":

        # При удалении объекта
        if 'delete_obj' in request.POST:
            # Если объект можно удалить - удаляем (всё связанное удалится каскадно)
            if can_be_deleted:
                object.delete()
                return HttpResponseRedirect('/workspace')
            # Иначе возвращаем сообщение об ошибке
            else:
                return HttpResponseForbidden('Этот объект нельзя удалить – он уже в работе')
        # При переводе объекта в состояние В сборке
        elif 'to_work_obj' in request.POST:
            # Если объект не В сборке - переводим его в это состояние
            if ready is False:
                # Удаляем текущее состояние По умолчанию (Приостановлен)
                obj_states = ObjectStateInstance.objects.filter(object=object)
                for state in obj_states:
                    if state.state == get_default_object_state():
                        state.delete()
                        break
                # Добавляем состояние В сборке
                ObjectStateInstance.objects.create(
                    object=object, state=get_ready_object_state(), created_at=timezone.now().date())
                ready = True
                context['ready'] = ready
            # Иначе возвращаем сообщение об ошибке
            else:
                return HttpResponseForbidden('Этот объект уже В сборке')
        # При переводе объекта в состояние Приостановлен
        elif 'stop_obj' in request.POST:
            # Если объект не Приостановлен - переводим его в это состояние
            if ready is True:
                # Удаляем текущее состояние В сборке
                obj_states = ObjectStateInstance.objects.filter(object=object)
                for state in obj_states:
                    if state.state == get_ready_object_state():
                        state.delete()
                        break
                # Добавляем состояние Приостановлен
                ObjectStateInstance.objects.create(
                    object=object, state=get_default_object_state(), created_at=timezone.now().date())
                ready = False
                context['ready'] = ready
            # Иначе возвращаем сообщение об ошибке
            else:
                return HttpResponseForbidden('Этот объект уже Приостановлен')
        # При скрытии объекта
        elif 'hide_obj' in request.POST:
            object.hidden = True
            object.save()
            context['object'] = object
        # При отображении объекта
        elif 'show_obj' in request.POST:
            object.hidden = False
            object.save()
            context['object'] = object

    # Если пришёл AJAX-запрос на обновление деталей объекта
    elif request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Проверяем кеш, чтобы не отправлять повторно одинаковые данные
        cache_key = f'object_detail_{object.id}'
        cache_data = []
        for product in products:
            cache_data.append(
                f'{product.id}: {product.get_ava_amount()}, {product.get_ava_parts_amount()}, {product.get_in_work_amount()}, {product.get_parts_in_work_amount()}')
        cur_hash = hashlib.md5(json.dumps(
            cache_data, sort_keys=True).encode()).hexdigest()
        prev_hash = cache.get(cache_key)
        if prev_hash and prev_hash == cur_hash:
            return JsonResponse({'html': ""})
        cache.set(cache_key, cur_hash, timeout=300)
        # Отправляем заполненный шаблон с деталями объекта
        data = {'html': render_to_string(
            "partials/object_details.html", context, request)}
        return JsonResponse(data)

    # Возвращаем заполненный шаблон с деталями объекта при стандартной прогрузке страницы
    return render(request, 'object_detail.html', context)


@login_required
def in_work_view(request: HttpRequest) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Страница со списком изделий, которые в данный момент изготавливаются всеми работниками (для мастера)
    ### Параметры
    - request — HTTP-запрос
    ### Возвращаемое значение
    - Заполненный шаблон со списком изделий в работе
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Если пришёл POST запрос (изделие отмечено как завершённое)
    if request.method == "POST":
        # Получаем запись о завершённом изделии
        finished_id = request.POST.get("work_id")
        instance = CreationInstance.objects.filter(id=finished_id).first()
        # Сбрасываем готовность объекта для последующего обновления
        object = instance.product.object if instance.product else instance.part.product.object
        object.ready_percentage = None
        object.save()
        # Сбрасываем доступное количество изделия/частей и завершённого количества изделия
        parts = Part.objects.filter(product=instance.product)
        for part in parts:
            part.ava_amount = None
            part.save()
        product = instance.product if instance.product else instance.part.product
        product.ava_amount = None
        product.completed_amount = None
        product.save()
        # Если запись о завершённом изделии/части уже есть, обновляем её (увеличилось кол-во завершённых изделий/частей)
        completed = CreationInstance.objects.filter(
            worker=instance.worker, product=instance.product, part=instance.part, status="COMPLETED").first()
        if completed:
            completed.amount += instance.amount
            completed.completed = timezone.now()
            completed.save()
            instance.delete()
        # Иначе помечаем текущее изделие/часть как завершённое
        else:
            instance.status = 'COMPLETED'
            instance.completed = timezone.now()
            instance.save()
        # Обновляем доступное и завершённое количество изделия/частей, готовность объекта
        for part in parts:
            part.get_ava_amount()
        product.get_ava_amount()
        product.get_completed_amount()
        object.get_ready_percentage()
    # Получаем все записи о изделиях/частях, которые в данный момент изготавливаются
    instances = (
        CreationInstance.objects
        .filter(status='IN_WORK')
        .select_related('product', 'part', 'part__product', 'worker')
    )
    questions_count = Question.objects.filter(answer='').count()
    # Отправляем заполненный шаблон
    context = {
        'instances': instances,
        'questions': questions_count
    }
    return render(request, 'in_work_list.html', context)


@login_required
def workers_list_view(request: HttpRequest) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Страница со списком всех работников и их статистикой (для мастера)
    ### Параметры
    - request — HTTP-запрос
    ### Возвращаемое значение
    - Заполненный шаблон со списком работников и их статистикой
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    context = dict()
    if request.method == "POST":
        # Если нужно добавить нового пользователя
        if 'add_user' in request.POST:
            # Создаём нужную для добавления пользователя форму
            form = CustomUserCreationForm()
            # Добавляем её в словарь с данными
            context['form'] = form
        # Если нужно сохранить созданного пользователя
        elif 'create_user' in request.POST:
            # Создаём форму для добавления пользователя и передаём в неё данные
            form = CustomUserCreationForm(request.POST)
            # Если форма корректна (нет ошибок в переданных значениях)
            if form.is_valid():
                # Сохраняем созданного пользователя (модель User)
                user = form.save()
                # Получаем из БД группу, соответствующую работнику
                group = Group.objects.get(name="worker")
                # Добавляем созданному пользователю группу "Работник"
                user.groups.add(group)
                # Вызываем метод проверки данных о работнике для их создания
                worker_data = check_worker_data(user=user)
                # Получаем отображаемое имя пользователя из формы
                display_name = form.cleaned_data["display_name"]
                # В данные о работнике добавляем отображаемое имя
                worker_data.display_name = display_name
                # Сохраняем данные о работнике
                worker_data.save()
            # Если произошла ошибка при заполнении формы
            # Возвращаем форму с сообщениями об ошибках
            else:
                context['form'] = form
    # Получаем из запроса дату (месяц, за который нужны данные)
    date = request.GET.get("date")
    # Если в запросе была дата
    if date:
        # Переводим её в подходящий формат
        cur_date = datetime.strptime(date, '%Y-%m-%d').date()
    # Если в запросе не было даты - используем текущую дату
    else:
        cur_date = datetime.now().date()
    # Определяем предыдущий месяц
    prev = cur_date - relativedelta(months=1)
    # Определяем следующий месяц
    next = cur_date + relativedelta(months=1)
    # Определяем начало выбранного месяца
    start = cur_date.replace(day=1)
    # Определяем конец выбранного месяца
    end = (start + relativedelta(months=1)
           ).replace(day=1) - relativedelta(days=1)
    # Получаем данные о всех работниках
    workers_data = WorkerData.objects.all().select_related('worker')
    month_creations = CreationInstance.objects.filter(
        completed__gte=start,
        completed__lte=end,
        status='COMPLETED'
    ).select_related('product', 'part')
    all_creations = CreationInstance.objects.filter(
        status='COMPLETED'
    ).select_related('product', 'part')
    from collections import defaultdict
    month_data_by_worker = defaultdict(lambda: {'completed': 0, 'payment': 0})
    for creation in month_creations:
        if creation.worker_id:
            month_data_by_worker[creation.worker_id]['completed'] += creation.amount
            month_data_by_worker[creation.worker_id]['payment'] += creation.get_price()

    total_data_by_worker = defaultdict(lambda: {'completed': 0, 'payment': 0})
    for creation in all_creations:
        if creation.worker_id:
            total_data_by_worker[creation.worker_id]['completed'] += creation.amount
            total_data_by_worker[creation.worker_id]['payment'] += creation.get_price()
    # Создаём пустой словарь для сбора данных о работниках
    workers = dict()
    all_workers_with_stats = []
    # Создаём переменную для хранения общего кол-ва произведённых изделий
    all_completed = 0
    # Создаём переменную для хранения общей суммы выплат
    all_payment = 0
    # Создаём переменную для хранения кол-ва произведённых изделий за выбранный месяц
    completed = 0
    # Создаём переменную для хранения суммы выплат за выбранный месяц
    payment = 0
    # Собираем данные о всех работниках
    for worker in workers_data:
        m_stats = month_data_by_worker.get(
            worker.id, {'completed': 0, 'payment': 0})
        t_stats = total_data_by_worker.get(
            worker.id, {'completed': 0, 'payment': 0})

        if m_stats['completed'] > 0:
            workers[worker] = {
                "worker": worker,
                "completed": m_stats['completed'],
                "payment": m_stats['payment']
            }
            completed += m_stats['completed']
            payment += m_stats['payment']

        all_completed += t_stats['completed']
        all_payment += t_stats['payment']

        all_workers_with_stats.append({
            'worker': worker,
            'all_completed': t_stats['completed'],
            'all_payment': t_stats['payment']
        })
    # Заполняем словарь с данными для шаблона
    context['workers'] = workers
    context['completed_products'] = completed
    context['payment'] = payment
    context['prev'] = prev
    context['next'] = next
    context['current_date'] = cur_date
    context['all_workers'] = all_workers_with_stats
    context['all_completed'] = all_completed
    context['all_payment'] = all_payment
    # Получаем кол-во вопросов, на которые не был дан ответ
    context['questions'] = len(Question.objects.filter(answer=''))
    # Возвращаем заполненный шаблон
    return render(request, 'workers_list.html', context)


@login_required
def product_in_work_detail_view(request: HttpRequest, pk: int) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Детальный обзор изделия для мастера
    ### Параметры
    - request — HTTP-запрос
    - pk — первичный ключ изделия (его id)
    ### Возвращаемое значение
    - Заполненный шаблон с деталями изделия
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем информацию о выбранном изделии
    product = get_object_or_404(Product, pk=pk)
    parts = Part.objects.filter(product=product)
    # Получаем все записи о данном изделии и его частях в работе, в очереди и завершённые
    in_work_products = CreationInstance.objects.filter(
        product=product, status='IN_WORK')
    queued_products = CreationInstance.objects.filter(
        product=product, status='QUEUED')
    completed_products = CreationInstance.objects.filter(
        product=product, status="COMPLETED")
    # Собираем в списки части изделия по их статусам
    in_work_parts = []
    queued_parts = []
    completed_parts = []
    for part in parts:
        raw_parts = CreationInstance.objects.filter(part=part)
        for raw_part in raw_parts:
            if raw_part.status == "IN_WORK":
                in_work_parts.append(raw_part)
            elif raw_part.status == "QUEUD":
                queued_parts.append(raw_part)
            elif raw_part.status == "COMPLETED":
                completed_parts.append(raw_part)
    # Заполняем контекст для шаблона
    context = {
        'product': product,
        'in_work_products': in_work_products,
        'queued_products': queued_products,
        'completed_products': completed_products,
        'in_work_parts': in_work_parts,
        'queued_parts': queued_parts,
        'completed_parts': completed_parts,
        'parts': parts
    }
    # Собираем части изделия, которые можно добавить в очередь
    raw_parts = Part.objects.filter(product=product)
    selectable_parts = None
    for part in raw_parts:
        if part.get_ava_amount() > 0:
            if selectable_parts is None:
                selectable_parts = [part]
            else:
                selectable_parts.append(part)
    # Собираем варианты выбора для формы добавления в очередь
    choices = None
    def_amount = 1
    def_choice = '1'
    if product.get_ava_amount() > 0:
        def_amount = min(def_amount, product.ava_float())
        choices = [('1', 'Всё изделие')]
    idx = 2
    if selectable_parts:
        for part in selectable_parts:
            if choices is None:
                def_amount = min(def_amount, part.get_ava_amount())
                def_choice = str(idx)
                choices = [(str(idx), part.name)]
                idx += 1
            else:
                choices.append((str(idx), part.name))
                idx += 1
    # Добавляем форму в контекст, если есть варианты выбора
    if choices:
        form = AddProductToQueueForm(choices=choices, initial={
            'amount': def_amount, 'creation': def_choice})
        context['queueform'] = form
    if request.method == "GET":
        # Режим редактирования описания изделия
        if 'edit' in request.GET:
            context['edit_mode'] = request.GET["edit"] == '1'
            form = EnterDescriptionForm(
                initial={'description': product.description})
            context['editform'] = form
        # При возврате к списку изделий
        elif 'return' in request.GET:
            return HttpResponseRedirect(f"/workspace/objects/{product.object.id}")
        # Если пришёл AJAX-запрос на обновление деталей изделия
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Проверяем кеш, чтобы не отправлять повторно одинаковые данные
            cache_key = f'product_detail_{product.id}'
            cache_data = [
                f'{product.id}: {product.get_ava_amount()}, {product.get_ava_parts_amount()}, {product.get_in_work_amount()}, {product.get_parts_in_work_amount()}']
            for part in parts:
                cache_data.append(
                    f'{part.id}: {part.get_ava_amount()}, {part.get_completed_amount()}, {part.get_in_work_amount()}')
            cur_hash = hashlib.md5(json.dumps(
                cache_data, sort_keys=True).encode()).hexdigest()
            prev_hash = cache.get(cache_key)
            if prev_hash and prev_hash == cur_hash:
                return JsonResponse({'html': ""})
            cache.set(cache_key, cur_hash, timeout=300)
            # Отправляем заполненный шаблон с деталями изделия
            data = {'html': render_to_string(
                "partials/product_in_work_details.html", context, request)}
            return JsonResponse(data)
    elif request.method == "POST":
        # Сохранение изменений описания изделия
        if 'save' in request.POST:
            form = EnterDescriptionForm(request.POST)
            if form.is_valid():
                description = form.cleaned_data["description"]
                product.description = description
                product.save()
                context['edit_mode'] = False
        # Отмена редактирования описания изделия
        elif 'cancel' in request.POST:
            context['edit_mode'] = False
        # Добавление изделия или его части в очередь на изготовление
        elif 'add_to_queue' in request.POST:
            # Передаём запрос в форму
            if choices:
                form = AddProductToQueueForm(request.POST, choices=choices, initial={
                    'amount': def_amount, 'creation': def_choice})
            else:
                form = AddProductToQueueForm(request.POST)
            # Добавляем изделие/часть в очередь работнику, если форма заполнена корректно
            if form.is_valid():
                # Получаем данные из формы
                amount = form.cleaned_data['amount']
                choice = int(form.cleaned_data['creation'][0])
                worker = form.cleaned_data['worker']
                worker_data = check_worker_data(user=worker)
                # Добавляем в очередь всё изделие
                if choice == 1:
                    if product.get_ava_amount() < amount:
                        form.add_error(
                            'amount', f'Выбрано недопустимое кол-во. К изготовлению доступно {product.get_ava_amount()} шт.')
                        context['queueform'] = form
                        return render(request, 'product_in_work.html', context)
                    # Сбрасываем доступное количество изделия для обновления
                    parts = Part.objects.filter(product=product)
                    for part in parts:
                        part.ava_amount = None
                        part.save()
                    # Сбрасываем доступное количество изделия
                    product.ava_amount = None
                    product.save()
                    # Если запись о данном изделии в очереди у данного работника уже есть - увеличиваем количество
                    instance = CreationInstance.objects.filter(
                        worker=worker_data, product=product, status='QUEUED').first()
                    if instance:
                        instance.amount += amount
                        instance.save()
                    # Иначе создаём новую запись
                    else:
                        CreationInstance.objects.create(
                            worker=worker_data, product=product, status='QUEUED', amount=amount, queued=timezone.now())
                    # Обновляем доступное количество частей
                    for part in parts:
                        part.get_ava_amount()
                    # Обновляем доступное количество изделия
                    product.get_ava_amount()
                # Добавляем в очередь часть изделия
                else:
                    selected_part = None
                    idx = 2
                    for part in selectable_parts:
                        if idx == choice:
                            selected_part = part
                            break
                        idx += 1
                    if amount > selected_part.get_ava_amount():
                        form.add_error(
                            'amount', f'Выбрано недопустимое кол-во. К изготовлению доступно {selected_part.get_ava_amount()} шт.')
                        context['queueform'] = form
                        return render(request, 'product_in_work.html', context)
                    selected_part.product.ava_amount = None
                    selected_part.product.save()
                    # Сбрасываем доступное количество частей для обновления
                    parts = Part.objects.filter(product=selected_part.product)
                    for part in parts:
                        part.ava_amount = None
                        part.save()
                    # Если запись о данной части в очереди у данного работника уже есть - увеличиваем количество
                    instance = CreationInstance.objects.filter(
                        worker=worker_data, part=selected_part, status='QUEUED').first()
                    if instance:
                        instance.amount += amount
                        instance.save()
                    # Иначе создаём новую запись
                    else:
                        CreationInstance.objects.create(
                            worker=worker_data, part=selected_part, status='QUEUED', amount=amount, queued=timezone.now())
                    for part in parts:
                        part.get_ava_amount()
                    # Обновляем доступное количество части и изделия
                    selected_part.product.get_ava_amount()
                # Обновляем список для добавления в очередь изделия/частей
                raw_parts = Part.objects.filter(product=product)
                selectable_parts = None
                for part in raw_parts:
                    if part.get_ava_amount() > 0:
                        if selectable_parts is None:
                            selectable_parts = [part]
                        else:
                            selectable_parts.append(part)
                choices = None
                def_amount = 1
                def_choice = '1'
                if product.get_ava_amount() > 0:
                    def_amount = min(def_amount, product.ava_float())
                    choices = [('1', 'Всё изделие')]
                idx = 2
                if selectable_parts:
                    for part in selectable_parts:
                        if choices is None:
                            def_amount = min(def_amount, part.get_ava_amount())
                            def_choice = str(idx)
                            choices = [(str(idx), part.name)]
                            idx += 1
                        else:
                            choices.append((str(idx), part.name))
                            idx += 1
                if choices:
                    form = AddProductToQueueForm(
                        choices=choices, initial={
                            'amount': def_amount, 'creation': def_choice})
                    context['queueform'] = form
                # Обновляем список частей в очереди
                queued_parts = []
                for part in parts:
                    queue_parts = CreationInstance.objects.filter(
                        part=part, status='QUEUED')
                    for queue_part in queue_parts:
                        queued_parts.append(queue_part)
                # Обновляем данные о изделиях/частях в очереди
                queued_products = CreationInstance.objects.filter(
                    product=product, status='QUEUED')
                context['queued_parts'] = queued_parts
                context['queued_products'] = queued_products
    # Возвращаем заполненный шаблон с деталями изделия
    return render(request, 'product_in_work.html', context)


@login_required
def worker_detail(request: HttpRequest, pk: int) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Детальный обзор работника для мастера
    ### Параметры
    - request — HTTP-запрос
    - pk — первичный ключ работника (его id)
    ### Возвращаемое значение
    - Заполненный шаблон с деталями работника
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем информацию о выбранном работнике
    worker_data = get_object_or_404(WorkerData, pk=pk)
    # Получаем дату из запроса
    date = request.GET.get("date")
    if date:
        cur_date = datetime.strptime(date, '%Y-%m-%d').date()
    else:
        cur_date = datetime.now().date()
    # Собираем предыдущий и следующий месяцы для навигации
    prev = cur_date - relativedelta(months=1)
    next = cur_date + relativedelta(months=1)
    # Собираем начало и конец месяца
    start = cur_date.replace(day=1)
    end = start
    while end.month == start.month:
        end += relativedelta(days=1)
    end -= relativedelta(days=1)
    # Получаем все завершённые изделия/части работника за выбранный месяц
    completed_products = CreationInstance.objects.filter(
        status="COMPLETED", completed__gte=start, completed__lte=end, worker=worker_data)
    # Получаем сумму выплат и кол-во завершённых изделий/частей за выбранный месяц
    payment = worker_data.get_payment(start, end)
    completed_amount = worker_data.get_completed(start, end)
    if request.method == "POST":
        # При удалении работника
        if 'delete_user' in request.POST:
            # Получаем ссылку на пользователя (класс User)
            worker = worker_data.worker
            # Открепляем данные о работнике от пользователя и удаляем пользователя
            worker_data.worker = None
            worker_data.save()
            worker.delete()
            # Перенаправляем на список работников
            return HttpResponseRedirect('/workspace/workers_list')
    # Получаем все изделия/части работника, которые в данный момент в работе или в очереди
    products_in_work = CreationInstance.objects.filter(
        worker=worker_data).filter(status__in=['IN_WORK', 'QUEUED'])
    # Заполняем контекст для шаблона
    context = {
        'worker': worker_data,
        'products': completed_products,
        'payment': payment,
        'completed_amount': completed_amount,
        'in_work': products_in_work,
        'current_date': cur_date,
        'prev': prev,
        'next': next
    }
    # Возвращаем заполненный шаблон с деталями работника
    return render(request, 'worker_detail.html', context)


@login_required
def questions_list(request: HttpRequest) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Страница со списком всех вопросов без ответов (для мастера)
    ### Параметры
    - request — HTTP-запрос
    ### Возвращаемое значение
    - Заполненный шаблон со списком вопросов без ответов
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем все вопросы, на которые не был дан ответ
    questions = Question.objects.filter(answer="")
    # Создаём словарь с данными для шаблона
    context = {
        'questions': questions,
    }
    # Возвращаем заполненный шаблон
    return render(request, 'questions_list.html', context)


@login_required
def instance_details(request: HttpRequest, pk: int) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Детальный обзор изделия/части для мастера
    ### Параметры
    - request — HTTP-запрос
    - pk — первичный ключ записи о создании изделия/части (его id)
    ### Возвращаемое значение
    - Заполненный шаблон с деталями изделия/части
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем информацию о выбранном изделии/части
    instance = get_object_or_404(CreationInstance, pk=pk)
    form = None
    if request.method == "GET":
        # Если нужно ввести ответ на вопрос
        if 'question' in request.GET:
            # Создаём форму для ввода ответа
            question_id = request.GET["question"]
            if Question.objects.filter(instance=instance, id=question_id).exists():
                form = EnterAnswerForm(
                    initial={'answer': Question.objects.filter(instance=instance, id=question_id).first().answer})
            else:
                return HttpResponseForbidden('Такого вопроса не существует')
    # Если нужно сохранить ответ на вопрос (пришёл POST-запрос)
    else:
        # Сохраянем введённый ответ
        question_id = request.GET["question"]
        if Question.objects.filter(instance=instance, id=question_id).exists():
            form = EnterAnswerForm(request.POST)
            if form.is_valid():
                answer = form.cleaned_data["answer"]
                question = Question.objects.filter(
                    instance=instance, id=question_id).first()
                question.answer = answer
                question.save()
        else:
            return HttpResponseForbidden('Такого вопроса не существует')
    # Получаем все вопросы, относящиеся к изделию/части
    questions = Question.objects.filter(instance=instance).all()
    # Заполняем контекст для шаблона
    context = {
        'instance': instance,
        'questions': questions,
        'form': form
    }
    # Возвращаем заполненный шаблон с деталями изделия/части
    return render(request, 'instance_detail.html', context)


def check_spec(workbook: xl.Workbook):
    """### Проверяет корректность формата спецификации

    Args:
        workbook (xl.Workbook): Excel-книга, содержащая спецификацию

    Raises:
        ValidationError: Причина несоответствия формату

    Returns:
        bool: Соответствует ли спецификация формату
    """
    # Проверка заголовков в спецификации
    required_headers = {
        'Наименование': False,
        'Кол-во': False,
        'ТЗ\nч/д': False,
        'Итого\nруб': False,
        'З/п': False
    }
    sheet = workbook['Спецификация']
    NAME_COLUMN = None
    first_row = next(sheet.iter_rows(
        min_row=1, max_row=1, values_only=False), [])
    for idx, cell in enumerate(first_row):
        val = cell.value
        if val in required_headers:
            required_headers[val] = True
            if val == 'Наименование':
                NAME_COLUMN = idx

    for header, located in required_headers.items():
        if not located:
            raise ValidationError(f'Не найден столбец {header}')

    # Цвет: #33CCFF
    # Закрашено и выделено жирным — заголовок изделия
    # Закрашено без выделения — заголовок части изделия
    rows_generator = sheet.iter_rows(min_row=11, values_only=False)
    # Флаг названия изделия
    isProduct = False
    # Флаг названия части
    isPart = False
    # Флаг наличия изделий
    anyProduct = False

    for row in rows_generator:
        cell = row[NAME_COLUMN]
        name = cell.value

        fill = cell.fill
        is_colored = fill and fill.start_color and fill.start_color.rgb == 'FF33CCFF'

        if is_colored:
            font = cell.font
            is_bold = font and font.bold

            if is_bold:
                if isProduct:
                    raise ValidationError('Обнаружено пустое изделие')
                isProduct = True
                anyProduct = True
            else:
                if isPart:
                    raise ValidationError('Обнаружена пустая часть')
                if not anyProduct:
                    raise ValidationError(
                        'Обнаружена часть, не принадлежащая изделию')
                isPart = True
                isProduct = False

        else:
            if not anyProduct:
                raise ValidationError('Обнаружено оборудование без изделия')
            isProduct = False
            isPart = False

    if not anyProduct:
        raise ValidationError('Не найдено ни одного изделия')

    return True


class ParseableProduct():

    def __init__(self, name: str, price: Decimal, labor_cost: Decimal, payment: Decimal):
        self.name = name
        self.number = ''
        self.price = price
        self.labor_cost = labor_cost
        self.payment = payment
        self.parts: list[ParseablePart] = []
        self.divIntoParts = True


class ParseablePart():

    def __init__(self, name: str, product_price: Decimal, product_payment: Decimal, products: list[ParseableProduct]):
        self.name = name
        self.price = Decimal(0)
        self.payment = Decimal(0)
        self.product_price = product_price
        self.product_payment = product_payment
        for product in products:
            product.parts.append(self)


def parse_spec(spec):
    """### Функция для парсинга спецификации
    Парсит изделия и части из спецификации в формате Excel-таблицы для дальнейшего переноса в БД

    Args:
        spec: Загруженная Excel-книга, содержащая спецификацию

    Returns:
        list[`class`:ParseableProduct:]: Список изделий
    """
    # Считываем данные из файла
    spec_format = xl.load_workbook(
        spec, read_only=True, data_only=True)
    sheet = spec_format['Спецификация']
    # Проверяем корректность файла
    check_spec(spec_format)
    # Находим положение нужных столбцов
    required_headers = {
        'Наименование': None,
        'Кол-во': None,
        'ТЗ\nч/д': None,
        'Итого\nруб': None,
        'З/п': None
    }
    first_row = next(sheet.iter_rows(
        min_row=1, max_row=1, values_only=False), [])
    for idx, cell in enumerate(first_row):
        if cell.value in required_headers:
            required_headers[cell.value] = idx
    NAME_COLUMN = required_headers['Наименование']
    AMOUNT_COLUMN = required_headers['Кол-во']
    LABOR_COSTS_COLUMN = required_headers['ТЗ\nч/д']
    PRICE_COLUMN = required_headers['Итого\nруб']
    PAYMENT_COLUMN = required_headers['З/п']

    # Парсим изделия и части
    blacklist_patterns = [p.value for p in ParseBlacklistValue.objects.all()]
    products: list[ParseableProduct] = []
    parent_products: list[ParseableProduct] = []
    current_part: ParseablePart | None = None
    D2 = Decimal('0.01')
    D4 = Decimal('0.0001')
    rows_generator = sheet.iter_rows(min_row=11, values_only=False)
    for row in rows_generator:
        name_cell = row[NAME_COLUMN]
        name = name_cell.value

        fill = name_cell.fill
        is_colored = fill and fill.start_color and fill.start_color.rgb == 'FF33CCFF'

        amount_val = row[AMOUNT_COLUMN].value
        amount = Decimal(amount_val if amount_val is not None else 1)

        price_val = row[PRICE_COLUMN].value
        price = Decimal(price_val if price_val is not None else 0).quantize(
            D2, rounding=ROUND_HALF_UP)

        # Парсим изделие или часть
        if is_colored:
            font = name_cell.font
            is_bold = font and font.bold

            # Изделие
            if is_bold:
                current_part = None
                parent_products = []

                labor_val = row[LABOR_COSTS_COLUMN].value
                labor_cost = Decimal(labor_val if labor_val is not None else 0).quantize(
                    D4, rounding=ROUND_HALF_UP)

                payment_val = row[PAYMENT_COLUMN].value
                payment = Decimal(payment_val if payment_val is not None else 0).quantize(
                    D2, rounding=ROUND_HALF_UP)

                if amount == 1:
                    products.append(ParseableProduct(
                        name, price, labor_cost, payment))
                    parent_products = [products[-1]]

                elif amount > 1:
                    products_list = name.split(', ')
                    for product_item in products_list:
                        if ' - ' in product_item:
                            start, end = product_item.split(' - ')
                            start_nums = [int(x)
                                          for x in re.findall(r'\d+', start)]
                            end_nums = [int(x)
                                        for x in re.findall(r'\d+', end)]

                            prefix_match = re.match(r'(^[^\d]+)', start)
                            prefix = prefix_match.group(
                                1) if prefix_match else ""

                            results = []
                            if start_nums[:-1] == end_nums[:-1]:
                                stable_nums = start_nums[:-1]
                                full_prefix = prefix + \
                                    ".".join(map(str, stable_nums)) + \
                                    "." if stable_nums else prefix

                                start_last = start_nums[-1]
                                end_last = end_nums[-1]
                                for last in range(start_last, end_last + 1):
                                    results.append(f"{full_prefix}{last}")

                            elif len(start_nums) == 2:
                                for major in range(start_nums[0], end_nums[0] + 1):
                                    s_minor = start_nums[1]
                                    e_minor = end_nums[1]
                                    for minor in range(s_minor, e_minor + 1):
                                        results.append(
                                            f"{prefix}{major}.{minor}")

                            for result in results:
                                products.append(ParseableProduct(
                                    result, price, labor_cost, payment))
                                parent_products.append(products[-1])
                        else:
                            products.append(ParseableProduct(
                                product_item, price, labor_cost, payment))
                            parent_products.append(products[-1])

                if len(parent_products) != amount:
                    raise ValidationError(
                        f'Количество сгенерированных наименований {name} не совпадает с заданным количеством ({len(parent_products)} != {amount})'
                    )

            # Часть
            else:
                current_part = ParseablePart(
                    name, parent_products[0].price, parent_products[0].payment, parent_products)
                if any(fnmatch(name, pattern) for pattern in blacklist_patterns):
                    for product in parent_products:
                        product.divIntoParts = False

        # Компонент
        else:
            if current_part:
                current_part.price += price
                current_part.payment = (current_part.product_payment * (
                    current_part.price / current_part.product_price)).quantize(D2, rounding=ROUND_HALF_UP)

    return products


@login_required
def migrate_view(request: HttpRequest) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Страница для импорта объекта из спецификации в формате Excel
    ### Параметры
    - request — HTTP-запрос
    ### Возвращаемое значение
    - Заполненный шаблон с формой для выбора файла спецификации
    - Перенаправление на главную страницу при отсутствии доступа
    - Список импортированных изделий и частей при успешном импорте
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    context = dict()
    # Получаем все вопросы, на которые не был дан ответ
    questions = Question.objects.filter(answer='')
    # Сохраняем кол-во вопросов в словарь данных для шаблона
    context['questions'] = len(questions)
    if request.method == "GET":
        form = SelectFileForm()
        context['form'] = form
    # Если нужно обработать файл
    elif request.method == "POST":
        form = SelectFileForm(request.POST, request.FILES)
        if form.is_valid():
            # Получаем файл из запроса
            spec = request.FILES.get("spec")
            obj_number = spec.name.split()[0]
            products = parse_spec(spec)

            # Добавляем записи в базу данных
            obj = Object.objects.create(obj_number=obj_number, created_at=timezone.now(
            ).date())
            ObjectStateInstance.objects.create(
                object=obj, state=get_default_object_state(), created_at=timezone.now())

            product_number = '1'
            number_len = len(str(len(products)))
            for product in products:
                if product.labor_cost > 0:
                    prod = Product.objects.create(prod_number=(number_len - len(product_number)) * '0' + product_number,
                                                  object=obj, name=product.name, amount=1, price=product.payment)
                    product.number = (
                        number_len - len(product_number)) * '0' + product_number
                    product_number = str(int(product_number)+1)
                    if product.divIntoParts:
                        for part in product.parts:
                            Part.objects.create(
                                name=part.name, product=prod, price=part.payment)

            context['products'] = [
                product for product in products if product.labor_cost > Decimal(0)]
            context['object'] = obj
    return render(request, "migrate.html", context)


@login_required
def queued_details(request: HttpRequest, pk: int) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Страница с подробностями изделия/части в очереди для работника
    ### Параметры
    - request — HTTP-запрос
    - pk — первичный ключ изделия/части (его id)
    ### Возвращаемое значение
    - Заполненный шаблон с деталями изделия/части в очереди
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "worker") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем данные о работнике из запроса
    worker = check_worker_data(request)
    # Получаем информацию об изделии/части в очереди
    instance = get_object_or_404(CreationInstance, pk=pk)
    # Проверяем, что у часть/изделие для данного работника
    if instance.worker != worker:
        return HttpResponseRedirect("/workspace")
    context = {'instance': instance}
    # Если пришёл запрос на взятие в работу изделия/части
    if request.method == "POST" and 'claim_product' in request.POST:
        # Если уже есть такое же изделие/часть в работе, увеличиваем его количество
        in_work = CreationInstance.objects.filter(
            worker=worker, product=instance.product, part=instance.part, status='IN_WORK').first()
        if in_work:
            in_work.amount += instance.amount
            in_work.save()
            instance.delete()
        # Иначе меняем статус изделия/части на "В работе"
        else:
            instance.status = "IN_WORK"
            instance.queued = None
            instance.started = timezone.now()
            instance.save()
        # После принятия изделия/части в работу перенаправляем на главную страницу
        return HttpResponseRedirect('/workspace')
    # Возвращаем заполненный шаблон с деталями изделия/части в очереди при стандартной загрузке
    return render(request, "queued.html", context)


@login_required
def hidden_view(request: HttpRequest) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Страница со списком всех скрытых объектов (для мастера)
    ### Параметры
    - request — HTTP-запрос
    ### Возвращаемое значение
    - Заполненный шаблон со списком скрытых объектов
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    # Получаем все скрытые объекты
    objects = Object.objects.filter(hidden=True)
    # Получаем все вопросы, на которые не был дан ответ
    questions = Question.objects.filter(answer='')
    # Получаем из запроса Поисковый запрос
    search_query = request.GET.get('search', '')
    # Если что-то было введено в поиск
    if search_query:
        # Оставляем только подходящие по номеру объекты
        objects = objects.filter(obj_number__icontains=search_query)
    # Создаём словарь с нужными данными
    context = {'objects': objects, 'hidden': True, 'questions': len(questions)}
    # Если пришёл запрос на динамическое обновление страницы
    # (Приходит после ввода в поисковое поле ИЛИ через определённый промежуток времени)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Возвращаем специальный шаблон, который будет включен в страницу
        return render(request, 'partials/objects_table.html', context)
    # Возвращаем заполненный шаблон страницы
    return render(request, 'master.html', context)


@login_required
def blacklist_settings_view(request: HttpRequest) -> HttpResponse | JsonResponse | HttpResponseRedirect:
    """
    ### Описание
    Страница с настройками черного списка парсинга (для мастера)
    ### Параметры
    - request — HTTP-запрос
    ### Возвращаемое значение
    - Заполненный шаблон с настройками черного списка парсинга
    - Перенаправление на главную страницу при отсутствии доступа
    """
    # Проверяем, что у пользователя есть доступ к этой странице
    if check_user_group(request, "master") is False:
        return HttpResponseRedirect('/workspace')
    # Обновляем уведомления (сработает, если пришёл AJAX-запрос)
    notify = update_notification(request)
    if notify:
        return notify
    blacklist = ParseBlacklistValue.objects.all()
    context = dict()
    # Получаем все вопросы, на которые не был дан ответ
    questions = Question.objects.filter(answer='')
    # Сохраняем кол-во вопросов в словарь данных для шаблона
    context['questions'] = len(questions)
    if request.method == "POST":
        # Если нужно добавить значение в черный список
        if 'add_value' in request.POST:
            # Получаем и добавляем значение из формы в черный список
            form = AddParseBlacklistValueForm(request.POST)
            if form.is_valid():
                value = form.cleaned_data["blacklist_value"]
                if value in blacklist:
                    form.add_error("blacklist_value",
                                   f"Маска {value} уже содержится в списке!")
                    context = {
                        "blacklist": blacklist,
                        "form": form,
                        'questions': len(questions),
                    }
                    return render(request, 'blacklist_settings.html', context)
                ParseBlacklistValue.objects.create(value=value)
                blacklist = ParseBlacklistValue.objects.all()
            context['form'] = form
        # Если нужно удалить значение из черного списка
        if 'delete' in request.POST:
            # Удаляем значение из черного списка согласно значению из формы
            ParseBlacklistValue.objects.filter(
                id=request.POST.get('delete')).first().delete()
            blacklist = ParseBlacklistValue.objects.all()
            form = AddParseBlacklistValueForm()
            context['form'] = form
    # При стандартной загрузке страницы создаём пустую форму для добавления значения в черный список
    else:
        form = AddParseBlacklistValueForm()
        context['form'] = form
    context['blacklist'] = blacklist
    # Возвращаем заполненный шаблон страницы
    return render(request, 'blacklist_settings.html', context)
